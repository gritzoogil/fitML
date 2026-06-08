"""
Import AUSNUT 2023 food data into PostgreSQL fitml database.
Run once: python import_ausnut.py
"""
import psycopg2
from openpyxl import load_workbook

DB_CONFIG = {
    'host': '127.0.0.1',
    'port': 5432,
    'dbname': 'fitml',
    'user': 'postgres',
    'password': ''
}

NUTRIENT_FILE = r'C:\Users\GilBryan\Downloads\AUSNUT 2023 - All Files_0\AUSNUT 2023 - All Files\AUSNUT 2023 - Food nutrient profiles.xlsx'
FOOD_FILE = r'C:\Users\GilBryan\Downloads\AUSNUT 2023 - All Files_0\AUSNUT 2023 - All Files\AUSNUT 2023 -Food details.xlsx'

def create_table(cur):
    cur.execute("""
        CREATE TABLE IF NOT EXISTS foods (
            id SERIAL PRIMARY KEY,
            food_key VARCHAR(20) UNIQUE NOT NULL,
            food_name TEXT NOT NULL,
            energy_kj DECIMAL(10,2),
            calories DECIMAL(10,2),
            protein_g DECIMAL(10,2),
            fat_g DECIMAL(10,2),
            carbs_g DECIMAL(10,2),
            sugar_g DECIMAL(10,2),
            fibre_g DECIMAL(10,2),
            sodium_mg DECIMAL(10,2),
            serving_size_g DECIMAL(10,2) DEFAULT 100,
            source VARCHAR(20) DEFAULT 'AUSNUT2023'
        );
        CREATE INDEX IF NOT EXISTS idx_foods_name ON foods USING gin(to_tsvector('english', food_name));
        CREATE INDEX IF NOT EXISTS idx_foods_name_trgm ON foods (food_name);
    """)
    print("Table created.")

def import_foods(cur):
    print("Loading nutrient profiles...")
    wb = load_workbook(NUTRIENT_FILE, read_only=True)
    ws = wb['Food nutrient profiles']
    
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=4, values_only=True)):
        if not row[1]:  # food_key
            continue
        
        food_key = str(row[1])
        food_name = str(row[3]) if row[3] else 'Unknown'
        energy_kj = float(row[4]) if row[4] else None
        # Convert kJ to kcal (1 kcal = 4.184 kJ)
        calories = round(float(row[4]) / 4.184, 1) if row[4] else None
        protein = float(row[7]) if row[7] else 0
        fat = float(row[8]) if row[8] else 0
        carbs = float(row[9]) if row[9] else 0
        sugar = float(row[12]) if row[12] else 0
        fibre = float(row[15]) if row[15] else 0
        sodium = float(row[25]) if row[25] else 0
        
        rows.append((food_key, food_name, energy_kj, calories, protein, fat, carbs, sugar, fibre, sodium))
    
    print(f"Inserting {len(rows)} foods...")
    cur.executemany("""
        INSERT INTO foods (food_key, food_name, energy_kj, calories, protein_g, fat_g, carbs_g, sugar_g, fibre_g, sodium_mg)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ON CONFLICT (food_key) DO UPDATE SET
            food_name = EXCLUDED.food_name,
            calories = EXCLUDED.calories,
            protein_g = EXCLUDED.protein_g,
            fat_g = EXCLUDED.fat_g,
            carbs_g = EXCLUDED.carbs_g
    """, rows)
    print(f"Inserted {len(rows)} foods successfully.")

def main():
    conn = psycopg2.connect(**DB_CONFIG)
    conn.autocommit = True
    cur = conn.cursor()
    
    create_table(cur)
    import_foods(cur)
    
    cur.execute("SELECT COUNT(*) FROM foods")
    count = cur.fetchone()[0]
    print(f"Total foods in database: {count}")
    
    cur.close()
    conn.close()
    print("Done!")

if __name__ == '__main__':
    main()